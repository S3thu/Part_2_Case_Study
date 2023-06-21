from openpyxl import load_workbook


from openpyxl import load_workbook


def update_excel(file_path, data):
    # Load the Excel file using openpyxl
    workbook = load_workbook(file_path)

    # Access the specific worksheet (assuming it's the first sheet)
    sheet = workbook.worksheets[0]

    # Update the necessary cells with the provided data
    sheet["A1"] = "Unread Messages"
    sheet["B1"] = "Unread Notifications"
    sheet["A2"] = data["messages_count"]
    sheet["B2"] = data["notifications_count"]

    # Save the changes to the Excel file
    workbook.save(file_path)
